#!/usr/bin/env python3
"""Streamlit frontend for broker payout pipeline."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import streamlit as st

APP_ROOT = Path(__file__).resolve().parent
RUNS_DIR = APP_ROOT / "streamlit_runs"
RUNS_DIR.mkdir(parents=True, exist_ok=True)

_pipeline_module = None


def get_pipeline():
    """Lazy load pipeline to avoid slow startup side-effects."""
    global _pipeline_module
    if _pipeline_module is None:
        import pipeline as pipeline_module
        _pipeline_module = pipeline_module
    return _pipeline_module


def input_kind(file_path: Path) -> str:
    return file_path.suffix.lower()


def save_uploaded_file(uploaded_file) -> Path:
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    run_dir = RUNS_DIR / run_id
    input_dir = run_dir / "input"
    input_dir.mkdir(parents=True, exist_ok=True)
    file_path = input_dir / Path(uploaded_file.name).name
    file_path.write_bytes(uploaded_file.getvalue())
    st.session_state["run_dir"] = run_dir
    return file_path


def file_details(file_path: Path) -> dict[str, Any]:
    kind = input_kind(file_path)
    if kind == ".csv":
        with file_path.open("r", encoding="utf-8", errors="ignore") as fh:
            lines = sum(1 for _ in fh)
        header = pd.read_csv(file_path, nrows=0)
        rows = max(lines - 1, 0)
        return {
            "sheet_count": 1,
            "sheets": [{"sheet": "CSV_DATA", "rows": rows, "cols": len(header.columns)}],
        }

    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets.append({"sheet": sheet_name, "rows": ws.max_row, "cols": ws.max_column})
    wb.close()
    return {"sheet_count": len(sheets), "sheets": sheets}


def preview_input(file_path: Path, sheet_name: str, nrows: int = 25) -> pd.DataFrame:
    if input_kind(file_path) == ".csv":
        return pd.read_csv(file_path, nrows=nrows)
    return pd.read_excel(file_path, sheet_name=sheet_name, nrows=nrows)


def run_pipeline(
    input_path: Path,
    output_dir: Path,
    engine: str,
    compare_engine: bool,
    effect_start: str,
    effect_end: str,
) -> dict[str, Any]:
    if input_kind(input_path) != ".xlsx":
        raise RuntimeError("Pipeline execution requires .xlsx input. CSV is preview/export only.")

    pipeline = get_pipeline()
    wb = openpyxl.load_workbook(str(input_path), data_only=True)
    sheets = pipeline.auto_detect_sheets(wb)
    if "rto_2w" not in sheets:
        raise RuntimeError("Could not find required 2W RTO sheet")

    selected_rows = []
    selected_hitl = None
    parity = None

    if compare_engine:
        pipeline.reset_hitl_queue()
        legacy_rows = pipeline.run_legacy_engine(wb, sheets, effect_start, effect_end)
        legacy_hitl = pipeline.snapshot_hitl_queue()

        pipeline.reset_hitl_queue()
        pandas_rows = pipeline.run_pandas_engine(str(input_path), sheets, effect_start, effect_end)
        pandas_hitl = pipeline.snapshot_hitl_queue()

        ok, messages = pipeline.compare_engine_outputs(legacy_rows, pandas_rows, legacy_hitl, pandas_hitl)
        parity = {"ok": ok, "messages": messages}
        if not ok:
            raise RuntimeError("Parity check failed. Disable compare mode or fix mismatches.")

        if engine == "legacy":
            selected_rows = legacy_rows
            selected_hitl = legacy_hitl
        else:
            selected_rows = pandas_rows
            selected_hitl = pandas_hitl
    else:
        pipeline.reset_hitl_queue()
        if engine == "legacy":
            selected_rows = pipeline.run_legacy_engine(wb, sheets, effect_start, effect_end)
        else:
            selected_rows = pipeline.run_pandas_engine(str(input_path), sheets, effect_start, effect_end)
        selected_hitl = pipeline.snapshot_hitl_queue()

    pipeline.restore_hitl_queue(selected_hitl)
    output_dir.mkdir(parents=True, exist_ok=True)
    file_groups = pipeline.route_and_write(selected_rows, str(output_dir))

    hitl_path = None
    if pipeline.hitl_queue.queue:
        hitl_path = output_dir / "hitl_review_queue.json"
        pipeline.hitl_queue.export_queue(str(hitl_path))

    manifest = []
    for output_file in sorted(output_dir.glob("*.xlsx")):
        wb_out = openpyxl.load_workbook(str(output_file), data_only=True, read_only=True)
        ws_out = wb_out.active
        manifest.append(
            {
                "name": output_file.name,
                "path": str(output_file),
                "rows": ws_out.max_row,
                "cols": ws_out.max_column,
                "size_kb": round(output_file.stat().st_size / 1024, 2),
            }
        )
        wb_out.close()

    return {
        "input_path": str(input_path),
        "output_dir": str(output_dir),
        "rows_generated": len(selected_rows),
        "files_written": len(file_groups),
        "hitl_count": len(pipeline.hitl_queue.queue),
        "hitl_path": str(hitl_path) if hitl_path else None,
        "manifest": manifest,
        "sheets_detected": sheets,
        "parity": parity,
        "effect_start": effect_start,
        "effect_end": effect_end,
        "engine": engine,
        "compare_engine": compare_engine,
    }


def main():
    st.set_page_config(page_title="Broker Payout Streamlit", layout="wide")
    st.title("Broker Payout Streamlit Console")
    st.caption("Upload workbook, run pipeline, preview input and output, and download files.")

    uploaded = st.file_uploader("Upload .xlsx or .csv", type=["xlsx", "csv"])
    if not uploaded:
        st.info("Upload a file to continue.")
        return

    uploaded_key = f"{uploaded.name}:{uploaded.size}"
    if st.session_state.get("active_file_key") != uploaded_key:
        st.session_state["active_file_key"] = uploaded_key
        st.session_state.pop("result", None)

    input_path = save_uploaded_file(uploaded)
    details = file_details(input_path)
    kind = input_kind(input_path)

    col_a, col_b, col_c = st.columns(3)
    col_a.metric("File", Path(input_path).name)
    col_b.metric("Size (KB)", round(input_path.stat().st_size / 1024, 2))
    col_c.metric("Sheets/Views", details["sheet_count"])

    st.subheader("Input Details")
    st.dataframe(pd.DataFrame(details["sheets"]), use_container_width=True)

    sheet_names = [s["sheet"] for s in details["sheets"]]
    input_sheet = st.selectbox("Preview input", sheet_names)
    st.dataframe(preview_input(input_path, input_sheet, nrows=25), use_container_width=True)

    pipeline = get_pipeline()
    auto_start, auto_end = pipeline.auto_detect_dates(str(input_path))

    st.subheader("Run Options")
    engine = st.selectbox("Engine", ["legacy", "pandas"], index=0)
    compare_engine = st.checkbox("Enable strict parity compare", value=False)
    use_auto_dates = st.checkbox("Use auto detected dates", value=True)

    if use_auto_dates:
        start_date, end_date = auto_start, auto_end
        st.caption(f"Auto period: {start_date} to {end_date}")
    else:
        col1, col2 = st.columns(2)
        start_date = str(col1.date_input("Start date", value=datetime.strptime(auto_start, "%Y-%m-%d").date()))
        end_date = str(col2.date_input("End date", value=datetime.strptime(auto_end, "%Y-%m-%d").date()))

    run_disabled = kind != ".xlsx"
    if run_disabled:
        st.warning("CSV is supported for preview. Pipeline run requires .xlsx input.")

    if st.button("Run Pipeline", type="primary", disabled=run_disabled):
        run_dir = st.session_state["run_dir"]
        output_dir = run_dir / "output"
        with st.spinner("Running pipeline..."):
            try:
                result = run_pipeline(
                    input_path=input_path,
                    output_dir=output_dir,
                    engine=engine,
                    compare_engine=compare_engine,
                    effect_start=start_date,
                    effect_end=end_date,
                )
                st.session_state["result"] = result
                st.success("Run completed")
            except Exception as exc:
                st.error(str(exc))

    result = st.session_state.get("result")
    if not result:
        return

    st.subheader("Run Summary")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Rows", result["rows_generated"])
    m2.metric("Files", result["files_written"])
    m3.metric("HITL", result["hitl_count"])
    m4.metric("Engine", result["engine"].upper())

    if result.get("parity"):
        if result["parity"]["ok"]:
            st.success("Parity check passed")
        else:
            st.error("Parity check failed")
        with st.expander("Parity diagnostics"):
            for line in result["parity"]["messages"]:
                st.write(f"- {line}")

    manifest_df = pd.DataFrame(result["manifest"])
    st.subheader("Generated Files")
    st.dataframe(manifest_df[["name", "rows", "cols", "size_kb"]], use_container_width=True)

    if result["manifest"]:
        names = [x["name"] for x in result["manifest"]]
        out_file = st.selectbox("Preview output file", names)
        out_meta = next(x for x in result["manifest"] if x["name"] == out_file)
        out_path = Path(out_meta["path"])

        preview_df = pd.read_excel(out_path, nrows=30)
        st.dataframe(preview_df, use_container_width=True)

        c1, c2 = st.columns(2)
        c1.download_button(
            label=f"Download {out_file} (.xlsx)",
            data=out_path.read_bytes(),
            file_name=out_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        csv_bytes = pd.read_excel(out_path).to_csv(index=False).encode("utf-8")
        c2.download_button(
            label=f"Download {Path(out_file).stem}.csv",
            data=csv_bytes,
            file_name=f"{Path(out_file).stem}.csv",
            mime="text/csv",
        )

    with st.expander("Run JSON"):
        st.code(json.dumps(result, indent=2), language="json")


if __name__ == "__main__":
    main()
