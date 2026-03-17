#!/usr/bin/env python3
"""
Broker Payout Pipeline — Automated CLI Tool
Transforms HM_DIGIT grid files into state-wise broker portal Excel files.

Usage:
    python pipeline.py input.xlsx                     # Auto-detect dates
    python pipeline.py input.xlsx -o results/         # Custom output dir
    python pipeline.py input.xlsx --start 2026-03-01 --end 2026-03-31
    python pipeline.py input.xlsx --dry-run            # Preview without writing
    python pipeline.py input.xlsx --engine pandas
    python pipeline.py input.xlsx --compare-engine --dry-run
"""

import argparse
import copy
import json
import openpyxl
from collections import Counter, defaultdict
from pathlib import Path
from datetime import datetime, timedelta
import re
import sys

try:
    import pandas as pd
except ImportError:
    pd = None


# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS (business rules — rarely change)
# ═══════════════════════════════════════════════════════════════════════════════

IC_CODE = "DIGIT"
PRODUCT_TYPE = "TW"
RULE_TYPE = "PAYINPAYOUT"
MISP_FALLBACK_PAYOUT = "2.5"

RTO_PREFIX_TO_STATE = {
    "AN": "ANDAMAN  ISLANDS", "AP": "ANDHRA PRADESH", "AR": "ARUNACHAL PRADESH",
    "AS": "ASSAM", "BR": "BIHAR", "CG": "CHHATTISGARH", "CH": "CHANDIGARH",
    "DD": "DADRA AND NAGAR HAVELI", "DL": "DELHI", "GA": "GOA", "GJ": "GUJARAT",
    "HP": "HIMACHAL PRADESH", "HR": "HARYANA", "JH": "JHARKHAND",
    "JK": "JAMMU AND KASHMIR", "KA": "KARNATAKA", "KL": "KERALA", "LA": "LADAKH",
    "MH": "MAHARASHTRA", "ML": "MEGHALAYA", "MN": "MANIPUR",
    "MP": "MADHYA PRADESH", "MZ": "MIZORAM", "NL": "NAGALAND", "OD": "ODISHA",
    "PB": "PUNJAB", "PY": "PUDUCHERRY", "RJ": "RAJASTHAN", "SK": "SIKKIM",
    "TN": "TAMIL NADU", "TR": "TRIPURA", "TS": "TELANGANA",
    "UK": "UTTARAKHAND", "UP": "UTTAR PRADESH", "WB": "WEST BENGAL",
}

STATE_FILE_GROUPS = {
    "ANDAMAN": ["AN"], "BIHAR": ["BR"], "DELHI": ["DL"], "GJ": ["GJ"],
    "GOA": ["GA"], "HP": ["HP"], "HR": ["HR"], "JH": ["JH"], "JK": ["JK"],
    "KARNATAK": ["KA"], "Madhyapradesh": ["MP", "CG"],
    "Maharashtra": ["MH"], "NORTHEAST": ["AR", "AS", "MN", "ML", "MZ", "NL", "SK", "TR"],
    "ORISSA": ["OD"], "PB_CH": ["PB", "CH"], "RJ": ["RJ"],
    "TN_KL": ["TN", "KL", "PY"], "UP_UK": ["UP", "UK"], "WB": ["WB"],
    "APTS": ["AP", "TS"],
}

PREFIX_TO_FILE_GROUP = {}
for _group, _prefixes in STATE_FILE_GROUPS.items():
    for _p in _prefixes:
        PREFIX_TO_FILE_GROUP[_p] = _group

OUTPUT_HEADERS = [
    "Rule Code", "Rule Name", "IC Code", "Product Type", "Group", "Rule Type",
    "Cover Type", "Business Type", "Vehicle Age", "State", "RTO",
    "Vehicle Category", "Vehicle Type", "Fuel Type", "Make", "Model",
    "Owner Type", "Usage Type", "Booking Mode", "Cover Selection Type", "Covers",
    "Addon Selection Type", "Addons", "CC From", "CC To", "Power From",
    "Power To", "GVW From", "GVW To", "Carrying From", "Carrying To",
    "NCB Type", "NCB From", "NCB To", "IDV From", "IDV To",
    "OD Discount From", "OD Discount To", "Effect Start Date", "Effect End Date",
    "PayIn (Commision Type)", "PayIn (Reward Type)", "PayIn (Amount Percentage)",
    "PayIn (OD Amount)", "PayIn (TP Amount)",
]


# ═══════════════════════════════════════════════════════════════════════════════
# AUTO-DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def auto_detect_sheets(wb):
    """
    Auto-detect which sheets are TW grids and which is the RTO lookup.
    Returns a dict of detected sheet roles.
    """
    sheets = {}
    all_names = wb.sheetnames

    # Find 2W RTO sheet
    for name in all_names:
        if "2W" in name.upper() and "RTO" in name.upper():
            sheets["rto_2w"] = name
            break

    # Find TW grid sheets by header patterns
    for name in all_names:
        ws = wb[name]
        # Check first 10 rows for header patterns
        headers = set()
        for r in range(1, min(6, ws.max_row + 1)):
            for c in range(1, min(15, ws.max_column + 1)):
                v = ws.cell(row=r, column=c).value
                if v:
                    headers.add(str(v).strip())

        name_upper = name.upper()

        # TW 1+1 & SATP: has "Agency/PB Clusters" + "CD1" + two "Max CD2"
        if ("1+1" in name_upper or "SATP" in name_upper) and "TW" in name_upper:
            sheets["tw_1plus1_satp"] = name

        # TW 1+5: has Make column + "1+5"
        elif "1+5" in name_upper and "TW" in name_upper:
            sheets["tw_1plus5"] = name

        # 2W Grid 5+5: has Make column + "5+5"
        elif "5+5" in name_upper and "2W" in name_upper:
            sheets["tw_5plus5"] = name

        # TW SAOD: has "SAOD"
        elif "SAOD" in name_upper and ("TW" in name_upper or "2W" in name_upper):
            sheets["tw_saod"] = name

    return sheets


def auto_detect_dates(source_path):
    """
    Auto-detect effect dates from the filename.
    E.g., 'HM_DIGIT_FEB26_GRID.xlsx' → Feb 2026
    Falls back to current month.
    """
    filename = Path(source_path).stem.upper()

    month_map = {
        "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
        "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
    }

    # Try to find month + year pattern like FEB26, MAR2026
    for month_str, month_num in month_map.items():
        pattern = rf"{month_str}[\s_-]*(\d{{2,4}})"
        match = re.search(pattern, filename)
        if match:
            year_str = match.group(1)
            year = int(year_str)
            if year < 100:
                year += 2000
            # Start of month
            start = datetime(year, month_num, 1)
            # End of month
            if month_num == 12:
                end = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end = datetime(year, month_num + 1, 1) - timedelta(days=1)
            return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")

    # Fallback: current month
    now = datetime.now()
    start = datetime(now.year, now.month, 1)
    if now.month == 12:
        end = datetime(now.year + 1, 1, 1) - timedelta(days=1)
    else:
        end = datetime(now.year, now.month + 1, 1) - timedelta(days=1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


from segment_parser import parse_segment, HITLQueue

# Global HITL queue — collects all low-confidence parsing results
hitl_queue = HITLQueue()


def reset_hitl_queue():
    """Reset global HITL queue between engine runs."""
    global hitl_queue
    hitl_queue = HITLQueue()


def snapshot_hitl_queue():
    """Capture HITL queue state for parity comparison."""
    return {
        "queue": copy.deepcopy(hitl_queue.queue),
        "auto_approved": hitl_queue.auto_approved,
        "manual_review": hitl_queue.manual_review,
    }


def restore_hitl_queue(snapshot):
    """Restore HITL queue state from a captured snapshot."""
    global hitl_queue
    hitl_queue = HITLQueue()
    hitl_queue.queue = copy.deepcopy(snapshot.get("queue", []))
    hitl_queue.auto_approved = snapshot.get("auto_approved", 0)
    hitl_queue.manual_review = snapshot.get("manual_review", 0)


def is_blank(value):
    """Treat None/NaN/empty strings as blank."""
    if value is None:
        return True
    if pd is not None and pd.isna(value):
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def to_clean_str(value):
    """Convert a scalar cell value to stripped string, preserving blanks as ''."""
    return "" if is_blank(value) else str(value).strip()


def parse_via_agent(segment_str, make_column=None, all_cluster_makes=None,
                    sheet_context="1plus1", row_context=None):
    """
    Parse a segment string via the AI Segment Parser.
    Returns list of pipeline-compatible dicts: [{vehicle_type, fuel, make, model, cc_from, cc_to}]
    Routes low-confidence results to the HITL queue.
    """
    results = parse_segment(
        segment_str,
        make_column=make_column,
        all_cluster_makes=all_cluster_makes,
        sheet_context=sheet_context,
    )

    output = []
    for parsed in results:
        # Route through HITL confidence gate
        hitl_queue.submit(parsed, row_context or {"segment": segment_str})

        output.append({
            "vehicle_type": parsed.to_portal_vehicle_type(),
            "fuel": parsed.to_portal_fuel(),
            "make": parsed.to_portal_make(),
            "model": "ALL",
            "cc_from": parsed.to_portal_cc_from(),
            "cc_to": parsed.to_portal_cc_to(),
        })

    return output


# ═══════════════════════════════════════════════════════════════════════════════
# VALUE TRANSFORM
# ═══════════════════════════════════════════════════════════════════════════════

def cd2_to_payout(value, skip_d=True):
    """Convert decimal CD2 to percentage string for PayIn column."""
    if is_blank(value):
        return None
    if isinstance(value, str):
        v = value.strip().upper()
        if v == "D":
            return None if skip_d else "0"
        if v == "MISP":
            return MISP_FALLBACK_PAYOUT
        try:
            value = float(v)
        except ValueError:
            return None

    pct = round(value * 100, 2)
    return str(int(pct)) if pct == int(pct) else str(pct)


# ═══════════════════════════════════════════════════════════════════════════════
# RTO LOOKUP
# ═══════════════════════════════════════════════════════════════════════════════

def build_rto_maps(wb, rto_sheet_name):
    """Build cluster → RTO list mappings (case-insensitive keys)."""
    ws = wb[rto_sheet_name]
    rto_map_c = defaultdict(set)
    rto_map_d = defaultdict(set)
    rto_map_e = defaultdict(set)

    for row_idx in range(3, ws.max_row + 1):
        rto_code = ws.cell(row=row_idx, column=2).value
        cluster_c = ws.cell(row=row_idx, column=3).value
        cluster_d = ws.cell(row=row_idx, column=4).value
        cluster_e = ws.cell(row=row_idx, column=5).value
        if not rto_code:
            continue
        rto_code = str(rto_code).strip()
        if cluster_c:
            rto_map_c[str(cluster_c).strip().lower()].add(rto_code)
        if cluster_d:
            rto_map_d[str(cluster_d).strip().lower()].add(rto_code)
        if cluster_e:
            rto_map_e[str(cluster_e).strip().lower()].add(rto_code)

    return (
        {k: sorted(v) for k, v in rto_map_c.items()},
        {k: sorted(v) for k, v in rto_map_d.items()},
        {k: sorted(v) for k, v in rto_map_e.items()},
    )


def get_rto_string(rto_list):
    return ", ".join(rto_list)

def get_state_from_rtos(rto_list):
    if not rto_list:
        return ""
    prefix = re.match(r"([A-Z]+)", rto_list[0])
    return RTO_PREFIX_TO_STATE.get(prefix.group(1), "") if prefix else ""

def make_rule_name(cluster, biz_type, cover_type):
    cluster_clean = cluster.upper().replace(" ", "_").replace(",", "").replace("+", "_")
    return f"{cluster_clean}_{biz_type.upper()}_{cover_type.upper()}"


# ═══════════════════════════════════════════════════════════════════════════════
# ROW BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def build_output_row(rule_name, cover_type, biz_type, vehicle_age,
                     state, rto_str, vehicle_type, fuel, make, model,
                     cc_from, cc_to, payout_str, effect_start, effect_end):
    return [
        None, rule_name, IC_CODE, PRODUCT_TYPE, None, RULE_TYPE,
        cover_type, biz_type, vehicle_age, state, rto_str,
        None, vehicle_type, fuel, make, model,
        "ALL", None, "any", "na", None, "na", None,
        cc_from, cc_to, None, None, None, None, None, None,
        "na", None, None, None, None, None, None,
        effect_start, effect_end,
        "net", "percentage", payout_str, None, None,
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# GRID PROCESSORS
# ═══════════════════════════════════════════════════════════════════════════════

def process_1plus1_satp(wb, sheet_name, rto_map, effect_start, effect_end):
    """Process TW 1+1 & SATP sheet → RR_COMP + ALL_TP."""
    ws = wb[sheet_name]
    all_rows = []

    grid_configs = [
        (5, "COMPREHENSIVE", "RR", "ALL"),
        (6, "TP", "ALL", "ALL"),
    ]

    for cd2_col, cover_type, biz_type, vehicle_age in grid_configs:
        cluster_rows = defaultdict(list)
        for row_idx in range(5, ws.max_row + 1):
            cluster = ws.cell(row=row_idx, column=2).value
            if not cluster or str(cluster).strip() in ("", " ", "Agency/PB Clusters"):
                continue
            cluster = str(cluster).strip()
            segment = ws.cell(row=row_idx, column=3).value
            cd2_val = ws.cell(row=row_idx, column=cd2_col).value
            if segment:
                cluster_rows[cluster].append((str(segment).strip(), cd2_val))

        for cluster, rows in cluster_rows.items():
            rto_list = rto_map.get(cluster.lower(), [])
            if not rto_list:
                continue
            rto_str = get_rto_string(rto_list)
            state = get_state_from_rtos(rto_list)
            rule_name = make_rule_name(cluster, biz_type, cover_type)

            for segment, cd2_val in rows:
                payout = cd2_to_payout(cd2_val, skip_d=True)
                if payout is None:
                    continue
                seg_rows = parse_via_agent(
                    segment, sheet_context="1plus1",
                    row_context={"cluster": cluster, "segment": segment, "sheet": sheet_name},
                )
                for seg_row in seg_rows:
                    all_rows.append(build_output_row(
                        rule_name, cover_type, biz_type, vehicle_age,
                        state, rto_str, seg_row["vehicle_type"], seg_row["fuel"],
                        seg_row["make"], seg_row.get("model", "ALL"),
                        seg_row["cc_from"], seg_row["cc_to"], payout,
                        effect_start, effect_end,
                    ))

    print(f"  [{sheet_name}] → {len(all_rows)} rows")
    return all_rows


def process_make_specific(wb, sheet_name, rto_map, cover_type, biz_type,
                          effect_start, effect_end, vehicle_age="ALL"):
    """Process TW 1+5 or 2W Grid 5+5 with EXCLUDE logic."""
    ws = wb[sheet_name]
    all_rows = []

    # Pass 1: Collect all explicit makes per cluster
    cluster_makes = defaultdict(set)
    for row_idx in range(3, ws.max_row + 1):
        cluster = ws.cell(row=row_idx, column=2).value
        if not cluster or str(cluster).strip() in ("", " ", "Agency/PB Clusters", "PB_Cluster", "All"):
            continue
        cluster = str(cluster).strip()
        make_raw = ws.cell(row=row_idx, column=3).value
        if make_raw:
            make_str = str(make_raw).strip()
            if make_str.upper() != "OTHERS":
                for m in re.split(r"[/,]", make_str):
                    m = m.strip()
                    if m:
                        cluster_makes[cluster].add(m.upper())

    # Pass 2: Generate output rows
    for row_idx in range(3, ws.max_row + 1):
        cluster = ws.cell(row=row_idx, column=2).value
        if not cluster or str(cluster).strip() in ("", " ", "Agency/PB Clusters", "PB_Cluster", "All"):
            continue
        cluster = str(cluster).strip()

        make_raw = ws.cell(row=row_idx, column=3).value
        segment = ws.cell(row=row_idx, column=4).value
        cd2_val = ws.cell(row=row_idx, column=6).value

        if not segment:
            continue

        make_str = str(make_raw).strip() if make_raw else "ALL"
        payout = cd2_to_payout(cd2_val, skip_d=False)
        if payout is None:
            continue

        rto_list = rto_map.get(cluster.lower(), [])
        if not rto_list:
            continue

        rto_str = get_rto_string(rto_list)
        state = get_state_from_rtos(rto_list)
        rule_name = make_rule_name(cluster, biz_type, cover_type)

        # EXCLUDE logic
        output_make = make_str
        if make_str.upper() == "OTHERS":
            explicit = cluster_makes.get(cluster, set())
            output_make = f"EXCLUDE: {', '.join(sorted(explicit))}" if explicit else "ALL"

        # Parse segment via AI agent (handles EXCLUDE, CC ranges, fuel)
        seg_variants = parse_via_agent(
            str(segment).strip(),
            make_column=make_str,
            all_cluster_makes=cluster_makes.get(cluster, set()),
            sheet_context="make_specific",
            row_context={"cluster": cluster, "make": make_str, "segment": str(segment).strip(), "sheet": sheet_name},
        )

        for p in seg_variants:
            # Override make: use agent-derived make for Others, else raw make
            final_make = p["make"] if make_str.upper() == "OTHERS" else output_make
            all_rows.append(build_output_row(
                rule_name, cover_type, biz_type, vehicle_age,
                state, rto_str, p["vehicle_type"], p["fuel"],
                final_make, "ALL", p.get("cc_from"), p.get("cc_to"),
                payout, effect_start, effect_end,
            ))

    print(f"  [{sheet_name}] → {len(all_rows)} rows")
    return all_rows


def process_saod(wb, sheet_name, rto_map, effect_start, effect_end):
    """Process TW SAOD sheet → multi-year SAOD rows."""
    ws = wb[sheet_name]
    all_rows = []
    year_configs = [(8, "1"), (9, "2"), (10, "3"), (11, "4")]

    for row_idx in range(6, ws.max_row + 1):
        cluster = ws.cell(row=row_idx, column=2).value
        if not cluster or str(cluster).strip() in ("", "Cluster"):
            continue
        cluster = str(cluster).strip()
        segment = ws.cell(row=row_idx, column=3).value
        if not segment:
            continue

        rto_list = rto_map.get(cluster.lower(), [])
        if not rto_list:
            continue

        rto_str = get_rto_string(rto_list)
        state = get_state_from_rtos(rto_list)
        seg_variants = parse_via_agent(
            str(segment).strip(), sheet_context="saod",
            row_context={"cluster": cluster, "segment": str(segment).strip(), "sheet": sheet_name},
        )
        if not seg_variants:
            continue

        for seg in seg_variants:
            for cd2_col, year_str in year_configs:
                cd2_val = ws.cell(row=row_idx, column=cd2_col).value
                payout = cd2_to_payout(cd2_val, skip_d=True)
                if payout is None:
                    continue
                rule_name = make_rule_name(cluster, "RR", "SAOD")
                all_rows.append(build_output_row(
                    rule_name, "SAOD", "RR", year_str,
                    state, rto_str, seg["vehicle_type"], seg["fuel"],
                    seg.get("make", "ALL"), "ALL",
                    seg.get("cc_from"), seg.get("cc_to"),
                    payout, effect_start, effect_end,
                ))

    print(f"  [{sheet_name}] → {len(all_rows)} rows")
    return all_rows


# ═══════════════════════════════════════════════════════════════════════════════
# PANDAS ENGINE (TRANSFORM LAYER)
# ═══════════════════════════════════════════════════════════════════════════════

def pd_load_sheet_df(source_path, sheet_name):
    """Load a sheet using pandas with raw header-less indexing."""
    if pd is None:
        raise RuntimeError("pandas is required for '--engine pandas' or '--compare-engine'.")
    return pd.read_excel(source_path, sheet_name=sheet_name, header=None, dtype=object)


def pd_build_rto_maps(rto_df):
    """Build cluster → RTO list mappings from pandas dataframe."""
    rto_map_c = defaultdict(set)
    rto_map_d = defaultdict(set)
    rto_map_e = defaultdict(set)

    for _, row in rto_df.iloc[2:].iterrows():
        rto_code = to_clean_str(row.get(1))
        cluster_c = to_clean_str(row.get(2))
        cluster_d = to_clean_str(row.get(3))
        cluster_e = to_clean_str(row.get(4))

        if not rto_code:
            continue
        if cluster_c:
            rto_map_c[cluster_c.lower()].add(rto_code)
        if cluster_d:
            rto_map_d[cluster_d.lower()].add(rto_code)
        if cluster_e:
            rto_map_e[cluster_e.lower()].add(rto_code)

    return (
        {k: sorted(v) for k, v in rto_map_c.items()},
        {k: sorted(v) for k, v in rto_map_d.items()},
        {k: sorted(v) for k, v in rto_map_e.items()},
    )


def pd_process_1plus1_satp(df, sheet_name, rto_map, effect_start, effect_end):
    """Pandas version of TW 1+1 & SATP processor."""
    all_rows = []
    grid_configs = [
        (4, "COMPREHENSIVE", "RR", "ALL"),  # col 5 (1-based)
        (5, "TP", "ALL", "ALL"),            # col 6 (1-based)
    ]

    for cd2_idx, cover_type, biz_type, vehicle_age in grid_configs:
        cluster_rows = defaultdict(list)
        for _, row in df.iloc[4:].iterrows():  # row 5 onward (1-based)
            cluster = to_clean_str(row.get(1))
            if cluster in ("", "Agency/PB Clusters"):
                continue
            segment = to_clean_str(row.get(2))
            cd2_val = row.get(cd2_idx)
            if segment:
                cluster_rows[cluster].append((segment, cd2_val))

        for cluster, rows in cluster_rows.items():
            rto_list = rto_map.get(cluster.lower(), [])
            if not rto_list:
                continue
            rto_str = get_rto_string(rto_list)
            state = get_state_from_rtos(rto_list)
            rule_name = make_rule_name(cluster, biz_type, cover_type)

            for segment, cd2_val in rows:
                payout = cd2_to_payout(cd2_val, skip_d=True)
                if payout is None:
                    continue
                seg_rows = parse_via_agent(
                    segment,
                    sheet_context="1plus1",
                    row_context={"cluster": cluster, "segment": segment, "sheet": sheet_name},
                )
                for seg_row in seg_rows:
                    all_rows.append(build_output_row(
                        rule_name, cover_type, biz_type, vehicle_age,
                        state, rto_str, seg_row["vehicle_type"], seg_row["fuel"],
                        seg_row["make"], seg_row.get("model", "ALL"),
                        seg_row["cc_from"], seg_row["cc_to"], payout,
                        effect_start, effect_end,
                    ))

    print(f"  [{sheet_name}] → {len(all_rows)} rows")
    return all_rows


def pd_process_make_specific(df, sheet_name, rto_map, cover_type, biz_type,
                             effect_start, effect_end, vehicle_age="ALL"):
    """Pandas version of make-specific processors (1+5, 5+5)."""
    all_rows = []
    cluster_makes = defaultdict(set)

    # Pass 1: collect explicit makes
    for _, row in df.iloc[2:].iterrows():  # row 3 onward (1-based)
        cluster = to_clean_str(row.get(1))
        if cluster in ("", "Agency/PB Clusters", "PB_Cluster", "All"):
            continue
        make_str = to_clean_str(row.get(2))
        if not make_str:
            continue
        if make_str.upper() != "OTHERS":
            for make_name in re.split(r"[/,]", make_str):
                make_name = make_name.strip()
                if make_name:
                    cluster_makes[cluster].add(make_name.upper())

    # Pass 2: generate output rows
    for _, row in df.iloc[2:].iterrows():
        cluster = to_clean_str(row.get(1))
        if cluster in ("", "Agency/PB Clusters", "PB_Cluster", "All"):
            continue

        make_raw = row.get(2)
        segment = to_clean_str(row.get(3))
        cd2_val = row.get(5)
        if not segment:
            continue

        make_str = to_clean_str(make_raw) if not is_blank(make_raw) else "ALL"
        payout = cd2_to_payout(cd2_val, skip_d=False)
        if payout is None:
            continue

        rto_list = rto_map.get(cluster.lower(), [])
        if not rto_list:
            continue

        rto_str = get_rto_string(rto_list)
        state = get_state_from_rtos(rto_list)
        rule_name = make_rule_name(cluster, biz_type, cover_type)

        output_make = make_str
        if make_str.upper() == "OTHERS":
            explicit = cluster_makes.get(cluster, set())
            output_make = f"EXCLUDE: {', '.join(sorted(explicit))}" if explicit else "ALL"

        seg_variants = parse_via_agent(
            segment,
            make_column=make_str,
            all_cluster_makes=cluster_makes.get(cluster, set()),
            sheet_context="make_specific",
            row_context={"cluster": cluster, "make": make_str, "segment": segment, "sheet": sheet_name},
        )

        for parsed in seg_variants:
            final_make = parsed["make"] if make_str.upper() == "OTHERS" else output_make
            all_rows.append(build_output_row(
                rule_name, cover_type, biz_type, vehicle_age,
                state, rto_str, parsed["vehicle_type"], parsed["fuel"],
                final_make, "ALL", parsed.get("cc_from"), parsed.get("cc_to"),
                payout, effect_start, effect_end,
            ))

    print(f"  [{sheet_name}] → {len(all_rows)} rows")
    return all_rows


def pd_process_saod(df, sheet_name, rto_map, effect_start, effect_end):
    """Pandas version of SAOD processor."""
    all_rows = []
    year_configs = [(7, "1"), (8, "2"), (9, "3"), (10, "4")]  # cols 8/9/10/11

    for _, row in df.iloc[5:].iterrows():  # row 6 onward (1-based)
        cluster = to_clean_str(row.get(1))
        if cluster in ("", "Cluster"):
            continue

        segment = to_clean_str(row.get(2))
        if not segment:
            continue

        rto_list = rto_map.get(cluster.lower(), [])
        if not rto_list:
            continue

        rto_str = get_rto_string(rto_list)
        state = get_state_from_rtos(rto_list)
        seg_variants = parse_via_agent(
            segment,
            sheet_context="saod",
            row_context={"cluster": cluster, "segment": segment, "sheet": sheet_name},
        )
        if not seg_variants:
            continue

        for seg in seg_variants:
            for cd2_idx, year_str in year_configs:
                cd2_val = row.get(cd2_idx)
                payout = cd2_to_payout(cd2_val, skip_d=True)
                if payout is None:
                    continue
                rule_name = make_rule_name(cluster, "RR", "SAOD")
                all_rows.append(build_output_row(
                    rule_name, "SAOD", "RR", year_str,
                    state, rto_str, seg["vehicle_type"], seg["fuel"],
                    seg.get("make", "ALL"), "ALL",
                    seg.get("cc_from"), seg.get("cc_to"),
                    payout, effect_start, effect_end,
                ))

    print(f"  [{sheet_name}] → {len(all_rows)} rows")
    return all_rows


def run_legacy_engine(wb, sheets, effect_start, effect_end):
    """Run original openpyxl transformation flow."""
    rto_map_c, rto_map_d, rto_map_e = build_rto_maps(wb, sheets["rto_2w"])
    print(f"    1+1/SATP clusters: {len(rto_map_c)}")
    print(f"    1+5/5+5 clusters:  {len(rto_map_d)}")
    print(f"    SAOD clusters:     {len(rto_map_e)}")

    print("\n[4/4] Processing grids...")
    all_rows = []
    if "tw_1plus1_satp" in sheets:
        all_rows.extend(process_1plus1_satp(
            wb, sheets["tw_1plus1_satp"], rto_map_c, effect_start, effect_end,
        ))
    if "tw_1plus5" in sheets:
        all_rows.extend(process_make_specific(
            wb, sheets["tw_1plus5"], rto_map_d, "COMPREHENSIVE", "NEW", effect_start, effect_end,
        ))
    if "tw_5plus5" in sheets:
        all_rows.extend(process_make_specific(
            wb, sheets["tw_5plus5"], rto_map_d, "COMPREHENSIVE", "RR", effect_start, effect_end,
        ))
    if "tw_saod" in sheets:
        all_rows.extend(process_saod(
            wb, sheets["tw_saod"], rto_map_e, effect_start, effect_end,
        ))
    return all_rows


def run_pandas_engine(source_path, sheets, effect_start, effect_end):
    """Run pandas-backed transformation flow with parity to legacy logic."""
    rto_df = pd_load_sheet_df(source_path, sheets["rto_2w"])
    rto_map_c, rto_map_d, rto_map_e = pd_build_rto_maps(rto_df)
    print(f"    1+1/SATP clusters: {len(rto_map_c)}")
    print(f"    1+5/5+5 clusters:  {len(rto_map_d)}")
    print(f"    SAOD clusters:     {len(rto_map_e)}")

    print("\n[4/4] Processing grids...")
    all_rows = []
    if "tw_1plus1_satp" in sheets:
        df_1plus1 = pd_load_sheet_df(source_path, sheets["tw_1plus1_satp"])
        all_rows.extend(pd_process_1plus1_satp(
            df_1plus1, sheets["tw_1plus1_satp"], rto_map_c, effect_start, effect_end,
        ))
    if "tw_1plus5" in sheets:
        df_1plus5 = pd_load_sheet_df(source_path, sheets["tw_1plus5"])
        all_rows.extend(pd_process_make_specific(
            df_1plus5, sheets["tw_1plus5"], rto_map_d, "COMPREHENSIVE", "NEW", effect_start, effect_end,
        ))
    if "tw_5plus5" in sheets:
        df_5plus5 = pd_load_sheet_df(source_path, sheets["tw_5plus5"])
        all_rows.extend(pd_process_make_specific(
            df_5plus5, sheets["tw_5plus5"], rto_map_d, "COMPREHENSIVE", "RR", effect_start, effect_end,
        ))
    if "tw_saod" in sheets:
        df_saod = pd_load_sheet_df(source_path, sheets["tw_saod"])
        all_rows.extend(pd_process_saod(
            df_saod, sheets["tw_saod"], rto_map_e, effect_start, effect_end,
        ))
    return all_rows


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT WRITER
# ═══════════════════════════════════════════════════════════════════════════════

def route_rows(all_rows):
    """Group rows by target state-file buckets without writing files."""
    file_groups = defaultdict(list)
    ungrouped = []

    for row in all_rows:
        rto_str = row[10]
        if not rto_str:
            ungrouped.append(row)
            continue
        first_rto = rto_str.split(",")[0].strip()
        prefix = re.match(r"([A-Z]+)", first_rto)
        if prefix:
            group = PREFIX_TO_FILE_GROUP.get(prefix.group(1))
            if group:
                file_groups[group].append(row)
            else:
                ungrouped.append(row)
        else:
            ungrouped.append(row)

    return file_groups, ungrouped


def route_and_write(all_rows, output_dir):
    """Group rows by state and write state-wise Excel files."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    file_groups, ungrouped = route_rows(all_rows)

    for group_name, rows in sorted(file_groups.items()):
        filename = f"{group_name}_2W_H&M.xlsx"
        filepath = output_path / filename
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Sheet1"
        ws_out.append(OUTPUT_HEADERS)
        for row in rows:
            ws_out.append(row)
        wb_out.save(filepath)
        print(f"    ✓ {filepath} ({len(rows)} rows)")

    if ungrouped:
        print(f"    ⚠ {len(ungrouped)} rows from multi-state REF clusters (unrouted)")

    return file_groups


# ═══════════════════════════════════════════════════════════════════════════════
# PARITY HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def normalize_row_for_compare(row):
    """Normalize row values for deterministic parity checks."""
    normalized = []
    for value in row:
        if is_blank(value):
            normalized.append(None)
        elif isinstance(value, str):
            normalized.append(value.strip())
        else:
            normalized.append(value)
    return tuple(normalized)


def rows_fingerprint(rows):
    """Create multiset fingerprint for order-insensitive row comparison."""
    return Counter(normalize_row_for_compare(row) for row in rows)


def normalize_hitl_item(item):
    """Normalize HITL queue item for parity comparison."""
    parsed = item.get("parsed", {})
    context = item.get("context", {})
    return (
        str(parsed.get("original_input", "")).strip(),
        str(parsed.get("canonical_form", "")).strip(),
        json.dumps(context, sort_keys=True),
    )


def compare_engine_outputs(legacy_rows, pandas_rows, legacy_hitl, pandas_hitl):
    """Compare legacy vs pandas outputs and HITL routing for strict parity."""
    messages = []
    ok = True

    if len(legacy_rows) != len(pandas_rows):
        ok = False
        messages.append(f"Row count mismatch: legacy={len(legacy_rows)} pandas={len(pandas_rows)}")
    else:
        messages.append(f"Row count parity OK: {len(legacy_rows)} rows")

    legacy_groups, legacy_ungrouped = route_rows(legacy_rows)
    pandas_groups, pandas_ungrouped = route_rows(pandas_rows)
    legacy_keys = set(legacy_groups.keys())
    pandas_keys = set(pandas_groups.keys())

    if legacy_keys != pandas_keys:
        ok = False
        messages.append(
            "File group mismatch: "
            f"legacy_only={sorted(legacy_keys - pandas_keys)}, "
            f"pandas_only={sorted(pandas_keys - legacy_keys)}"
        )
    else:
        messages.append(f"File groups parity OK: {len(legacy_keys)} groups")

    if len(legacy_ungrouped) != len(pandas_ungrouped):
        ok = False
        messages.append(
            f"Ungrouped row mismatch: legacy={len(legacy_ungrouped)} pandas={len(pandas_ungrouped)}"
        )

    for group in sorted(legacy_keys | pandas_keys):
        legacy_group_rows = legacy_groups.get(group, [])
        pandas_group_rows = pandas_groups.get(group, [])

        if len(legacy_group_rows) != len(pandas_group_rows):
            ok = False
            messages.append(
                f"[{group}] count mismatch: legacy={len(legacy_group_rows)} "
                f"pandas={len(pandas_group_rows)}"
            )
            continue

        legacy_norm = [normalize_row_for_compare(row) for row in legacy_group_rows]
        pandas_norm = [normalize_row_for_compare(row) for row in pandas_group_rows]
        if legacy_norm == pandas_norm:
            messages.append(f"[{group}] exact order parity OK ({len(legacy_group_rows)} rows)")
            continue

        legacy_fp = rows_fingerprint(legacy_group_rows)
        pandas_fp = rows_fingerprint(pandas_group_rows)
        if legacy_fp == pandas_fp:
            messages.append(f"[{group}] content parity OK (row order differs)")
        else:
            ok = False
            messages.append(f"[{group}] row-content mismatch")
            legacy_only = list((legacy_fp - pandas_fp).keys())
            pandas_only = list((pandas_fp - legacy_fp).keys())
            if legacy_only:
                messages.append(f"[{group}] sample legacy-only row: {legacy_only[0]}")
            if pandas_only:
                messages.append(f"[{group}] sample pandas-only row: {pandas_only[0]}")

    if legacy_hitl.get("auto_approved") != pandas_hitl.get("auto_approved"):
        ok = False
        messages.append(
            "HITL auto-approved mismatch: "
            f"legacy={legacy_hitl.get('auto_approved')} pandas={pandas_hitl.get('auto_approved')}"
        )
    if legacy_hitl.get("manual_review") != pandas_hitl.get("manual_review"):
        ok = False
        messages.append(
            "HITL manual-review mismatch: "
            f"legacy={legacy_hitl.get('manual_review')} pandas={pandas_hitl.get('manual_review')}"
        )

    legacy_hitl_counter = Counter(normalize_hitl_item(item) for item in legacy_hitl.get("queue", []))
    pandas_hitl_counter = Counter(normalize_hitl_item(item) for item in pandas_hitl.get("queue", []))
    if legacy_hitl_counter != pandas_hitl_counter:
        ok = False
        messages.append("HITL queue item mismatch")
    else:
        messages.append(f"HITL queue parity OK ({len(legacy_hitl.get('queue', []))} items)")

    return ok, messages

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Broker Payout Pipeline — Transform grid files into broker portal Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python pipeline.py HM_DIGIT_FEB26_GRID.xlsx
  python pipeline.py grid.xlsx -o results/ --start 2026-03-01 --end 2026-03-31
  python pipeline.py grid.xlsx --dry-run
        """,
    )
    parser.add_argument("input", help="Path to the HM_DIGIT grid Excel file")
    parser.add_argument("-o", "--output", default="pipeline_output",
                        help="Output directory (default: pipeline_output/)")
    parser.add_argument("--start", help="Effect start date (YYYY-MM-DD). Auto-detected from filename if omitted.")
    parser.add_argument("--end", help="Effect end date (YYYY-MM-DD). Auto-detected from filename if omitted.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview row counts without writing files")
    parser.add_argument(
        "--engine",
        choices=["legacy", "pandas"],
        default="legacy",
        help="Transformation engine: legacy (openpyxl) or pandas (default: legacy).",
    )
    parser.add_argument(
        "--compare-engine",
        action="store_true",
        help="Run both engines and fail if parity checks do not match.",
    )

    args = parser.parse_args()

    source_path = Path(args.input)
    if not source_path.exists():
        print(f"✗ File not found: {args.input}")
        sys.exit(1)
    if (args.engine == "pandas" or args.compare_engine) and pd is None:
        print("✗ pandas is not installed. Install dependencies and retry.")
        sys.exit(1)

    # ── Display config ──
    print("=" * 60)
    print("  BROKER PAYOUT PIPELINE")
    print("=" * 60)
    print(f"  Input:  {source_path}")
    print(f"  Output: {args.output}/")
    print(f"  Engine: {args.engine}")
    if args.compare_engine:
        print("  Compare: enabled (legacy vs pandas)")

    # ── Auto-detect dates ──
    if args.start and args.end:
        effect_start, effect_end = args.start, args.end
    else:
        effect_start, effect_end = auto_detect_dates(str(source_path))
        print(f"  Dates:  {effect_start} → {effect_end}  (auto-detected from filename)")

    if args.start:
        effect_start = args.start
    if args.end:
        effect_end = args.end
    print(f"  Period: {effect_start} to {effect_end}")

    # ── Load workbook ──
    print(f"\n[1/4] Loading workbook...")
    wb = openpyxl.load_workbook(str(source_path), data_only=True)

    # ── Auto-detect sheets ──
    print(f"[2/4] Auto-detecting sheet roles...")
    sheets = auto_detect_sheets(wb)
    for role, name in sheets.items():
        print(f"    {role:20s} → '{name}'")

    if "rto_2w" not in sheets:
        print("✗ Could not find 2W RTO sheet. Aborting.")
        sys.exit(1)

    # ── Build and process via selected engine(s) ──
    print(f"\n[3/4] Building RTO lookup maps...")
    all_rows = []

    if args.compare_engine:
        print("\n  Running legacy engine for parity baseline...")
        reset_hitl_queue()
        legacy_rows = run_legacy_engine(wb, sheets, effect_start, effect_end)
        legacy_hitl = snapshot_hitl_queue()

        print("\n  Running pandas engine for parity check...")
        reset_hitl_queue()
        pandas_rows = run_pandas_engine(str(source_path), sheets, effect_start, effect_end)
        pandas_hitl = snapshot_hitl_queue()

        ok, messages = compare_engine_outputs(legacy_rows, pandas_rows, legacy_hitl, pandas_hitl)
        print("\n  Engine parity report:")
        for msg in messages:
            print(f"    - {msg}")

        if not ok:
            print("\n✗ Engine parity check failed.")
            sys.exit(2)
        print("\n✓ Engine parity check passed.")

        if args.engine == "legacy":
            all_rows = legacy_rows
            restore_hitl_queue(legacy_hitl)
        else:
            all_rows = pandas_rows
            restore_hitl_queue(pandas_hitl)
    else:
        reset_hitl_queue()
        if args.engine == "legacy":
            all_rows = run_legacy_engine(wb, sheets, effect_start, effect_end)
        else:
            all_rows = run_pandas_engine(str(source_path), sheets, effect_start, effect_end)

    print(f"\n  Total rows generated: {len(all_rows)}")

    # ── Write output ──
    if args.dry_run:
        print(f"\n  [DRY RUN] Would write {len(all_rows)} rows to {args.output}/")
    else:
        print(f"\n  Writing to {args.output}/...")
        file_groups = route_and_write(all_rows, args.output)

        # ── HITL Confidence Report ──
        print(f"\n  Segment Parser Confidence Report:")
        print(hitl_queue.summary())
        if hitl_queue.queue:
            hitl_path = Path(args.output) / "hitl_review_queue.json"
            hitl_queue.export_queue(str(hitl_path))
            print(f"  ⚠ Exported {len(hitl_queue.queue)} items to {hitl_path}")

        print(f"\n{'=' * 60}")
        print(f"  ✓ DONE — {len(file_groups)} files written to '{args.output}/'")
        print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
