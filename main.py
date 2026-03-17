#!/usr/bin/env python3
"""FastAPI frontend for broker payout pipeline runs."""

from __future__ import annotations

import json
import mimetypes
import os
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any

# If user runs `streamlit run main.py`, delegate to Streamlit UI.
if os.environ.get("STREAMLIT_SERVER_PORT"):
    from streamlit_app import main as streamlit_main
    streamlit_main()
    raise SystemExit(0)

import openpyxl
import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

APP_ROOT = Path(__file__).resolve().parent
RUNS_DIR = APP_ROOT / "frontend_runs"
TEMPLATES_DIR = APP_ROOT / "templates"
STATIC_DIR = APP_ROOT / "static"

RUNS_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
STATIC_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="Broker Payout UI", version="1.0.0")
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
_pipeline_module = None


def _pipeline():
    """Lazy-load pipeline module to avoid expensive import side-effects at app startup."""
    global _pipeline_module
    if _pipeline_module is None:
        import pipeline as pipeline_module
        _pipeline_module = pipeline_module
    return _pipeline_module


def _safe_name(name: str) -> str:
    return Path(name).name.replace(" ", "_")


def _file_kind(file_path: Path) -> str:
    return file_path.suffix.lower()


def _save_upload(run_dir: Path, upload: UploadFile) -> Path:
    input_dir = run_dir / "input"
    input_dir.mkdir(parents=True, exist_ok=True)
    target = input_dir / _safe_name(upload.filename or "input.xlsx")
    data = upload.file.read()
    target.write_bytes(data)
    return target


def _csv_shape(file_path: Path) -> tuple[int, int]:
    with file_path.open("r", encoding="utf-8", errors="ignore") as fh:
        line_count = sum(1 for _ in fh)
    try:
        header = pd.read_csv(file_path, nrows=0)
    except (pd.errors.EmptyDataError, pd.errors.ParserError, UnicodeDecodeError):
        return 0, 0
    rows = max(line_count - 1, 0)
    return rows, len(header.columns)


def _input_details(file_path: Path) -> dict[str, Any]:
    if _file_kind(file_path) == ".csv":
        rows, cols = _csv_shape(file_path)
        return {
            "sheet_count": 1,
            "sheets": [{"name": "CSV_DATA", "rows": rows, "cols": cols}],
        }

    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheets.append({"name": sheet_name, "rows": ws.max_row, "cols": ws.max_column})
    wb.close()
    return {"sheet_count": len(sheets), "sheets": sheets}


def _input_preview_html(file_path: Path, sheet_name: str, nrows: int = 20) -> str:
    try:
        if _file_kind(file_path) == ".csv":
            preview_df = pd.read_csv(file_path, nrows=nrows)
        else:
            preview_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=nrows)
    except Exception as exc:
        return f"<p>Input preview unavailable: {escape(str(exc))}</p>"
    preview_df = preview_df.fillna("")
    return preview_df.to_html(index=False, classes="preview-table", border=0)


def _output_manifest(output_dir: Path) -> list[dict[str, Any]]:
    rows = []
    output_files = sorted(list(output_dir.glob("*.xlsx")) + list(output_dir.glob("*.csv")))
    for output_file in output_files:
        if _file_kind(output_file) == ".csv":
            csv_rows, csv_cols = _csv_shape(output_file)
            rows.append(
                {
                    "name": output_file.name,
                    "size_kb": round(output_file.stat().st_size / 1024, 2),
                    "rows": csv_rows,
                    "cols": csv_cols,
                }
            )
            continue

        wb = openpyxl.load_workbook(str(output_file), data_only=True, read_only=True)
        ws = wb.active
        rows.append(
            {
                "name": output_file.name,
                "size_kb": round(output_file.stat().st_size / 1024, 2),
                "rows": ws.max_row,
                "cols": ws.max_column,
            }
        )
        wb.close()
    return rows


def _output_preview_html(output_path: Path, nrows: int = 30) -> str:
    try:
        if _file_kind(output_path) == ".csv":
            df = pd.read_csv(output_path, nrows=nrows).fillna("")
        else:
            df = pd.read_excel(output_path, nrows=nrows).fillna("")
    except Exception as exc:
        return f"<p>Output preview unavailable: {escape(str(exc))}</p>"
    return df.to_html(index=False, classes="preview-table", border=0)


def _run_pipeline(
    source_path: Path,
    output_dir: Path,
    engine: str,
    compare_engine: bool,
    effect_start: str,
    effect_end: str,
) -> dict[str, Any]:
    p = _pipeline()
    wb = openpyxl.load_workbook(str(source_path), data_only=True)
    sheets = p.auto_detect_sheets(wb)
    if "rto_2w" not in sheets:
        raise RuntimeError("Could not detect required sheet: 2W RTO")

    parity = None
    chosen_rows = []
    chosen_hitl = None

    if compare_engine:
        p.reset_hitl_queue()
        legacy_rows = p.run_legacy_engine(wb, sheets, effect_start, effect_end)
        legacy_hitl = p.snapshot_hitl_queue()

        p.reset_hitl_queue()
        pandas_rows = p.run_pandas_engine(str(source_path), sheets, effect_start, effect_end)
        pandas_hitl = p.snapshot_hitl_queue()

        parity_ok, parity_messages = p.compare_engine_outputs(
            legacy_rows, pandas_rows, legacy_hitl, pandas_hitl
        )
        parity = {"ok": parity_ok, "messages": parity_messages}

        if engine == "legacy":
            chosen_rows, chosen_hitl = legacy_rows, legacy_hitl
        else:
            chosen_rows, chosen_hitl = pandas_rows, pandas_hitl
    else:
        p.reset_hitl_queue()
        if engine == "legacy":
            chosen_rows = p.run_legacy_engine(wb, sheets, effect_start, effect_end)
        else:
            chosen_rows = p.run_pandas_engine(str(source_path), sheets, effect_start, effect_end)
        chosen_hitl = p.snapshot_hitl_queue()

    p.restore_hitl_queue(chosen_hitl)
    output_dir.mkdir(parents=True, exist_ok=True)
    file_groups = p.route_and_write(chosen_rows, str(output_dir))

    hitl_path = None
    if p.hitl_queue.queue:
        hitl_path = output_dir / "hitl_review_queue.json"
        p.hitl_queue.export_queue(str(hitl_path))

    return {
        "rows_generated": len(chosen_rows),
        "files_written": len(file_groups),
        "hitl_count": len(p.hitl_queue.queue),
        "hitl_path": str(hitl_path) if hitl_path else None,
        "sheets_detected": sheets,
        "parity": parity,
    }


def _load_run_meta(run_id: str) -> dict[str, Any]:
    meta_path = RUNS_DIR / run_id / "run_meta.json"
    if not meta_path.exists():
        raise HTTPException(status_code=404, detail="Run not found")
    return json.loads(meta_path.read_text())


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "run": None,
            "error": None,
        },
    )


@app.post("/run", response_class=HTMLResponse)
async def run_job(
    request: Request,
    file: UploadFile = File(...),
    engine: str = Form("legacy"),
    compare_engine: str | None = Form(None),
    start_date: str = Form(""),
    end_date: str = Form(""),
):
    if not file.filename:
        return templates.TemplateResponse(
            "index.html",
            {"request": request, "run": None, "error": "Please upload a valid .xlsx or .csv file."},
            status_code=400,
        )
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in {".xlsx", ".csv"}:
        return templates.TemplateResponse(
            "index.html",
            {"request": request, "run": None, "error": "Only .xlsx and .csv uploads are supported."},
            status_code=400,
        )

    compare = compare_engine == "on"
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    run_dir = RUNS_DIR / run_id
    input_path = _save_upload(run_dir, file)
    output_dir = run_dir / "output"

    if file_ext == ".csv":
        details = _input_details(input_path)
        meta = {
            "run_id": run_id,
            "engine": engine,
            "compare_engine": compare,
            "effect_start": "",
            "effect_end": "",
            "input_path": str(input_path),
            "output_dir": str(output_dir),
            "input_details": details,
            "summary": None,
            "notice": "CSV uploaded successfully. Preview is available; pipeline execution requires .xlsx.",
            "error": None,
        }
        (run_dir / "run_meta.json").write_text(json.dumps(meta, indent=2))
        return RedirectResponse(url=f"/run/{run_id}", status_code=303)

    p = _pipeline()
    auto_start, auto_end = p.auto_detect_dates(str(input_path))
    effect_start = start_date.strip() or auto_start
    effect_end = end_date.strip() or auto_end

    try:
        summary = _run_pipeline(
            source_path=input_path,
            output_dir=output_dir,
            engine=engine,
            compare_engine=compare,
            effect_start=effect_start,
            effect_end=effect_end,
        )
    except Exception as exc:
        details = _input_details(input_path)
        meta = {
            "run_id": run_id,
            "engine": engine,
            "compare_engine": compare,
            "effect_start": effect_start,
            "effect_end": effect_end,
            "input_path": str(input_path),
            "output_dir": str(output_dir),
            "input_details": details,
            "summary": None,
            "notice": None,
            "error": str(exc),
        }
        (run_dir / "run_meta.json").write_text(json.dumps(meta, indent=2))
        return RedirectResponse(url=f"/run/{run_id}", status_code=303)

    details = _input_details(input_path)
    meta = {
        "run_id": run_id,
        "engine": engine,
        "compare_engine": compare,
        "effect_start": effect_start,
        "effect_end": effect_end,
        "input_path": str(input_path),
        "output_dir": str(output_dir),
        "input_details": details,
        "summary": summary,
        "notice": None,
        "error": None,
    }
    (run_dir / "run_meta.json").write_text(json.dumps(meta, indent=2))

    return RedirectResponse(url=f"/run/{run_id}", status_code=303)


@app.get("/run/{run_id}", response_class=HTMLResponse)
def view_run(request: Request, run_id: str, input_sheet: str | None = None, output_file: str | None = None):
    meta = _load_run_meta(run_id)
    input_path = Path(meta["input_path"])
    output_dir = Path(meta["output_dir"])

    input_sheets = [sheet["name"] for sheet in meta["input_details"]["sheets"]]
    selected_input_sheet = input_sheet if input_sheet in input_sheets else (input_sheets[0] if input_sheets else None)
    input_preview_html = (
        _input_preview_html(input_path, selected_input_sheet) if selected_input_sheet else "<p>No view available.</p>"
    )

    manifest = _output_manifest(output_dir) if output_dir.exists() else []
    output_names = [row["name"] for row in manifest]
    selected_output = output_file if output_file in output_names else (output_names[0] if output_names else None)
    output_preview_html = (
        _output_preview_html(output_dir / selected_output) if selected_output else "<p>No output file generated.</p>"
    )

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "run": meta,
            "manifest": manifest,
            "selected_input_sheet": selected_input_sheet,
            "selected_output": selected_output,
            "input_preview_html": input_preview_html,
            "output_preview_html": output_preview_html,
            "error": meta.get("error"),
        },
    )


@app.get("/download/{run_id}/{file_name}")
def download_output(run_id: str, file_name: str):
    file_path = RUNS_DIR / run_id / "output" / _safe_name(file_name)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    media_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    return FileResponse(
        path=str(file_path),
        filename=file_path.name,
        media_type=media_type,
    )


if __name__ == "__main__":
    import sys
    import uvicorn

    # Prevent accidental `streamlit run main.py` from blocking forever.
    if os.environ.get("STREAMLIT_SERVER_PORT"):
        print("This file is the FastAPI app. Run: streamlit run streamlit_app.py")
        sys.exit(0)

    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
